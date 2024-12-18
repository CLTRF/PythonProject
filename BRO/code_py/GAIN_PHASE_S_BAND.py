'''
Author: CLT, date_ 12/12/2024
'this function measures the S2P parameter for the S Band LNA
'''

import sys, os

#from Tools.scripts.patchcheck import status
#from pyvisa.constants import VI_ERROR_BERR, VI_ATTR_WIN_BASE_ADDR_32, VI_ATTR_WIN_BASE_ADDR_64

sys.path.insert(1, 'C:/Users/CLT/PycharmProjects/PythonProject/BRO/Include')

import class_VNA as vna
import pickle

def build_file_name(_LNA_number, _LNA_serial_number):

    warehouse_directory = 'H:\\DATA_WARE_HOUSE'
    if not os.path.exists(warehouse_directory):
        os.makedirs(warehouse_directory)

    warehouse_sub_directory = warehouse_directory + '\\data'
    if not os.path.exists(warehouse_sub_directory):
        os.makedirs(warehouse_sub_directory)

    if (_LNA_number == 0):
        warehouse_sub_directory = warehouse_directory + '\\data\\CAL'
        if not os.path.exists(warehouse_sub_directory):
            os.makedirs(warehouse_sub_directory)
    else:
        # warehouse_sub_directory = warehouse_directory + '\\data\\LNA'+str(_LNA_number)
        warehouse_sub_directory = warehouse_directory + '\\data\\LNA' + str(_LNA_number) + '\\' + 'SN' + str(
            _LNA_serial_number)
        if not os.path.exists(warehouse_sub_directory):
            os.makedirs(warehouse_sub_directory)
        if not os.path.exists(warehouse_sub_directory):
            os.makedirs(warehouse_sub_directory)

    if (_LNA_number == 0):
        object_name = 'cal_kit'
    else:
        object_name = 'LNA' + str(_LNA_number)

    warehouse_file_name = warehouse_sub_directory + '\\' + object_name + '.pkl'
    # warehouse_file_name             =   'H:\DATA_WARE_HOUSE\VNA_calibration.pkl'

    return warehouse_file_name, warehouse_sub_directory


def read_pkl_object( file_name):
    VNA_TEMP = vna.VNA()
    with open(file_name, 'rb') as input:
        VNA_TEMP = pickle.load(input)
    return VNA_TEMP

def Plot_and_Save_Delta_Gain(Frequency_Vector, Phase_Table, Index_Ref, File_name_fig, _item_1_name,_item_2_name, _item_3_name, _item_4_name, limit_positive, limit_negative):
    '''
     Standard plotting functions for the Unwrap phase object

     Arguments:
         Phase_Table : numpy type of.

             - 'XZ' = ['theta', [0]] - sweep angle theta at phi 0 degrees.
             - 'YZ' = ['theta', [90]] - sweep angle theta at phi 90 degrees.
             - 'XY' = ['phi', [90]] - sweep angle phi at theta 90 degrees.

         xtick_spacing (int): defines the x axis major ticks. Since these plots are angles this\
             effectively sets the angular resolution of the grid of the plot.

         linear_plot_offset (int): defined the degrees of offset to be applied to the data along\
             the **X axis** in case of linear plots. This is particular useful for linear plots\
             of gain in case the antenna primary gain is in the theta=0 degrees direction -\
             typical patch antenna. Setting this to 180 degrees brings the peak at the center \
             of the plot.

     Returns:
         object: returns a handle to the newly created figure. It also outputs the figure in\
             PDF/SVG and PNG formats at the output folder in the settings.
     '''

    import sys
    import os
    import numpy as np
    from mpl_toolkits.mplot3d import Axes3D
    from matplotlib import cm
    import matplotlib.pyplot as plt
    from matplotlib.colors import BoundaryNorm
    from matplotlib.ticker import MaxNLocator

    _Vector_Phase_Reference = Phase_Table[Index_Ref]
    _Vector_Diff_LNA_1 = Phase_Table[0] - Phase_Table[Index_Ref]
    _Vector_Diff_LNA_2 = Phase_Table[1] - Phase_Table[Index_Ref]
    _Vector_Diff_LNA_3 = Phase_Table[2] - Phase_Table[Index_Ref]
    _Vector_Diff_LNA_4 = Phase_Table[3] - Phase_Table[Index_Ref]

    data_to_plot = []
    plot_axis = Frequency_Vector
    fig = plt.figure()
    plt.title('Phase unwrap differences_reference_LNA:'+str(Index_Ref))
    plt.legend('Phase unwrap differences_reference_LNA:'+str(Index_Ref))
    plt.xlabel('Frequency in GHz')
    plt.ylabel('Delta angle in degrees')

    ax1 = fig.add_axes([0.1, 0.1, 0.8, 0.9])

    ax1.plot(plot_axis, _Vector_Diff_LNA_1)
    ax1.plot(plot_axis, _Vector_Diff_LNA_2)
    ax1.plot(plot_axis, _Vector_Diff_LNA_3)
    ax1.plot(plot_axis, _Vector_Diff_LNA_4)

    plt.grid(True)

    plt.savefig(File_name_fig+_item_1_name+_item_2_name+_item_3_name+_item_4_name+'.png')
    plt.savefig(File_name_fig+_item_1_name+_item_2_name+_item_3_name+_item_4_name+'.pdf')
    plt.savefig(File_name_fig+_item_1_name+_item_2_name+_item_3_name+_item_4_name+'.eps')

    ## plt.show()

    return fig


def Plot_and_Save_Delta_Phase(Frequency_Vector, Phase_Table, Index_Ref, File_name_fig, _item_1_name,_item_2_name, _item_3_name, _item_4_name, limit_positive, limit_negative):
    '''
     Standard plotting functions for the Unwrap phase object

     Arguments:
         Phase_Table : numpy type of.

             - 'XZ' = ['theta', [0]] - sweep angle theta at phi 0 degrees.
             - 'YZ' = ['theta', [90]] - sweep angle theta at phi 90 degrees.
             - 'XY' = ['phi', [90]] - sweep angle phi at theta 90 degrees.

         xtick_spacing (int): defines the x axis major ticks. Since these plots are angles this\
             effectively sets the angular resolution of the grid of the plot.

         linear_plot_offset (int): defined the degrees of offset to be applied to the data along\
             the **X axis** in case of linear plots. This is particular useful for linear plots\
             of gain in case the antenna primary gain is in the theta=0 degrees direction -\
             typical patch antenna. Setting this to 180 degrees brings the peak at the center \
             of the plot.

     Returns:
         object: returns a handle to the newly created figure. It also outputs the figure in\
             PDF/SVG and PNG formats at the output folder in the settings.
     '''

    import sys
    import os
    import numpy as np
    from mpl_toolkits.mplot3d import Axes3D
    from matplotlib import cm
    import matplotlib.pyplot as plt
    from matplotlib.colors import BoundaryNorm
    from matplotlib.ticker import MaxNLocator

    _Vector_Phase_Reference = Phase_Table[Index_Ref]
    _Vector_Diff_LNA_1 = Phase_Table[0] - Phase_Table[Index_Ref]
    _Vector_Diff_LNA_2 = Phase_Table[1] - Phase_Table[Index_Ref]
    _Vector_Diff_LNA_3 = Phase_Table[2] - Phase_Table[Index_Ref]
    _Vector_Diff_LNA_4 = Phase_Table[3] - Phase_Table[Index_Ref]

    data_to_plot = []
    plot_axis = Frequency_Vector
    fig = plt.figure()
    plt.title('Phase unwrap differences_reference_LNA:'+str(Index_Ref))
    plt.legend('Phase unwrap differences_reference_LNA:'+str(Index_Ref))
    plt.xlabel('Frequency in GHz')
    plt.ylabel('Delta angle in degrees')

    #ax1 = fig.add_axes([0.1, 0.1, 0.8, 0.9])

    plt.plot(plot_axis, _Vector_Diff_LNA_1)
    plt.plot(plot_axis, _Vector_Diff_LNA_2)
    plt.plot(plot_axis, _Vector_Diff_LNA_3)
    plt.plot(plot_axis, _Vector_Diff_LNA_4)

    plt.axhline(y=limit_positive, color='r', linestyle='dotted')
    plt.axhline(y=limit_negative, color='r', linestyle='dotted')

    plt.grid(True)


    plt.legend([_item_1_name,_item_2_name, _item_3_name, _item_4_name,'upper limit','lower limit'], loc="upper left")

    plt.savefig(File_name_fig+_item_1_name+_item_2_name+_item_3_name+_item_4_name+'.png')
    plt.savefig(File_name_fig+_item_1_name+_item_2_name+_item_3_name+_item_4_name+'.pdf')
    plt.savefig(File_name_fig+_item_1_name+_item_2_name+_item_3_name+_item_4_name+'.eps')

    return fig

def Plot_and_Save_Magnitude_Phase(Frequency_Vector, Phase_Table, File_name_fig, _item_1_name,_item_2_name, _item_3_name, _item_4_name):
    '''
     Standard plotting functions for the Unwrap phase object

     Arguments:
         Phase_Table : numpy type of.

             - 'XZ' = ['theta', [0]] - sweep angle theta at phi 0 degrees.
             - 'YZ' = ['theta', [90]] - sweep angle theta at phi 90 degrees.
             - 'XY' = ['phi', [90]] - sweep angle phi at theta 90 degrees.

         xtick_spacing (int): defines the x axis major ticks. Since these plots are angles this\
             effectively sets the angular resolution of the grid of the plot.

         linear_plot_offset (int): defined the degrees of offset to be applied to the data along\
             the **X axis** in case of linear plots. This is particular useful for linear plots\
             of gain in case the antenna primary gain is in the theta=0 degrees direction -\
             typical patch antenna. Setting this to 180 degrees brings the peak at the center \
             of the plot.

     Returns:
         object: returns a handle to the newly created figure. It also outputs the figure in\
             PDF/SVG and PNG formats at the output folder in the settings.
     '''

    import sys
    import os
    import numpy as np
    from mpl_toolkits.mplot3d import Axes3D
    from matplotlib import cm
    import matplotlib.pyplot as plt
    from matplotlib.colors import BoundaryNorm
    from matplotlib.ticker import MaxNLocator

    _Vector_Diff_LNA_1 = Phase_Table[0]
    _Vector_Diff_LNA_2 = Phase_Table[1]
    _Vector_Diff_LNA_3 = Phase_Table[2]
    _Vector_Diff_LNA_4 = Phase_Table[3]

    data_to_plot = []
    plot_axis = Frequency_Vector
    fig = plt.figure()
    plt.xlabel('Frequency in GHz')
    plt.ylabel('Phase angle in degrees')

    #ax1 = fig.add_axes([0.1, 0.1, 0.8, 0.9])

    plt.plot(plot_axis, _Vector_Diff_LNA_1)
    plt.plot(plot_axis, _Vector_Diff_LNA_2,)
    plt.plot(plot_axis, _Vector_Diff_LNA_3,)
    plt.plot(plot_axis, _Vector_Diff_LNA_4,)

    plt.legend([_item_1_name,_item_2_name, _item_3_name, _item_4_name], loc="lower right")

    plt.grid(True)

    plt.savefig(File_name_fig+_item_1_name+_item_2_name+_item_3_name+_item_4_name+'.png')
    plt.savefig(File_name_fig+_item_1_name+_item_2_name+_item_3_name+_item_4_name+'.pdf')
    plt.savefig(File_name_fig+_item_1_name+_item_2_name+_item_3_name+_item_4_name+'.eps')

    ## plt.show()

    return fig


def _S_BAND_SPARAMETER(_LNA_no, _Serial_Number,  _str_IP_vector_analyzer):

    import sys,os

#sys.path.insert(1, '/home/max/python/myimport')
    sys.path.insert(1, 'C:/Users/CLT/PycharmProjects/PythonProject/BRO/Include')

    import csv
    import copy
    import logging
    import time
    import datetime
    from math import inf
    from types import SimpleNamespace
    import numpy as np
    import gs_instrument
    from gs_instrument import network_analyzer
    import os

    _S2P_file_name          =   []
    _Gain_Phase_Parameter   =   []
    warehouse_file_name = 'H:\\DATA_WARE_HOUSE\\'

    if not os.path.exists('H:\\DATA_WARE_HOUSE' + '\\' + 'data\\'):
        os.makedirs('H:\\DATA_WARE_HOUSE' + '\\' + 'data\\')


    _ZNB20=network_analyzer.RohdeSchwarzZNB8(_str_IP_vector_analyzer)

    time.sleep(5)
    _now = datetime.datetime.now()
    _now_string = _now.strftime("%d%m%y_%H%M%S")

    if (_LNA_no > 0):
        _file_name_for_saving_sc = warehouse_file_name + 'data' + "\\" + "LNA" + str(_LNA_no) + '\\'+'SN'+str(_Serial_Number)+ "\\" + "LNA"+str(_LNA_no)+"_" + _now_string +"_Gain_"+"Screenshot.png"
        _file_name_for_saving_S2P = warehouse_file_name + 'data' + "\\" + "LNA" + str(_LNA_no) + "\\"+'SN'+str(_Serial_Number)+ "\\" +"LNA"+str(_LNA_no)+"_" + _now_string +"Gain.s2p"
    else:
        _file_name_for_saving_sc = warehouse_file_name + 'data' + "\\" + "CAL" + "\\" + "LNA" + str(
            _LNA_no) + "_" + _now_string + "_Gain_" + "Screenshot.png"
        _file_name_for_saving_S2P = warehouse_file_name + 'data' + "\\" + "CAL" + "\\" + "LNA" + str(
            _LNA_no) + "_" + _now_string + "Gain.s2p"


    _ZNB20.save_screenshot_png(_file_name_for_saving_sc)
    _data = _ZNB20.save_touchstone(_file_name_for_saving_S2P, trace =1 )

    return (True, _file_name_for_saving_sc,_file_name_for_saving_S2P,_data)


def save_to_excel(patterns_obj, filename = 'Default_XLS'):

    from openpyxl import load_workbook
    from openpyxl import Workbook
    import csv
    import copy
    import logging
    import time
    import datetime
    from math import inf
    from types import SimpleNamespace
    import numpy as np
    import gs_instrument
    from gs_instrument import network_analyzer

    _S2P_file_name          =   []
    _Gain_Phase_Parameter   =   []

    '''
    Saves selected patterns to an Excel file

    Args:
        patterns_obj (object): object of class Patterns. See\
                :class:`~includes.patterns`
        pat_inds (list): List of integers of the patterns to be stored
        field (str): string describing the field to be stored. See\
                :func:`~includes.field_operations.convert_field`
        field_format (str): field format of the field to be stored. See\
                :func:`~includes.field_operations.convert_field_format`
        filename (str): Filename to use. The file will be stored in the output folder. If a file of\
            the same name exists new sheets will be added. If the file does not exist it will be\
            created.
    '''
    if not os.path.exists('output'):
        os.makedirs('output')
    filename = 'output//'+filename+'.xlsx'
    if os.path.isfile(filename):
        workbook = load_workbook(filename)
    else:
        workbook = Workbook()

##    temp_field,_,context = patterns_obj.(pat, field, field_format)

    worksheet = workbook.create_sheet(title=str(pat)+'_'+field+'_'+field_format)
##    __fill_worksheet(worksheet, context, temp_field)   .Gain
    workbook.save(filename)